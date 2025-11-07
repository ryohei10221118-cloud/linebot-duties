"""
班表解析程序
用于解析 Excel 格式的班表并统计各员工的班别情况
"""

import openpyxl
from datetime import datetime
from shift_classifier import classify_shift, analyze_shifts


class ScheduleParser:
    """班表解析器"""

    def __init__(self, excel_path, sheet_name=None):
        """
        初始化解析器

        参数:
            excel_path (str): Excel 文件路径
            sheet_name (str, optional): 工作表名称，默认使用第一个工作表
        """
        self.workbook = openpyxl.load_workbook(excel_path)
        if sheet_name:
            self.sheet = self.workbook[sheet_name]
        else:
            self.sheet = self.workbook.active

    def parse_schedule(self, start_row=1, name_col=2):
        """
        解析班表数据

        参数:
            start_row (int): 数据开始的行号（1-based）
            name_col (int): 姓名所在的列号（1-based）

        返回:
            dict: 员工班表数据
        """
        schedule_data = {}

        # 读取日期行（假设在第1行）
        dates = []
        for col in range(name_col + 1, self.sheet.max_column + 1):
            cell_value = self.sheet.cell(row=1, column=col).value
            if cell_value:
                dates.append(cell_value)

        # 遍历所有员工行
        for row in range(start_row, self.sheet.max_row + 1):
            name_cell = self.sheet.cell(row=row, column=name_col)
            employee_name = name_cell.value

            # 跳过空行或标题行
            if not employee_name or isinstance(employee_name, (int, float)):
                continue

            # 跳过汇总行
            if '限休人数' in str(employee_name) or '已休数目' in str(employee_name):
                continue

            employee_name = str(employee_name).strip()

            # 读取该员工的所有班别
            shifts = []
            for col in range(name_col + 1, min(name_col + 1 + len(dates), self.sheet.max_column + 1)):
                shift_cell = self.sheet.cell(row=row, column=col)
                shift_value = shift_cell.value
                if shift_value:
                    shifts.append(str(shift_value).strip())
                else:
                    shifts.append("")

            schedule_data[employee_name] = {
                'shifts': shifts,
                'dates': dates[:len(shifts)]
            }

        return schedule_data

    def analyze_employee_schedule(self, employee_name, schedule_data):
        """
        分析单个员工的班表

        参数:
            employee_name (str): 员工姓名
            schedule_data (dict): 班表数据

        返回:
            dict: 分析结果
        """
        if employee_name not in schedule_data:
            return None

        employee_shifts = schedule_data[employee_name]['shifts']
        # 过滤掉空班别
        valid_shifts = [s for s in employee_shifts if s]

        stats = analyze_shifts(valid_shifts)

        return {
            'name': employee_name,
            'total_days': len(valid_shifts),
            'stats': stats,
            'shift_details': list(zip(schedule_data[employee_name]['dates'], employee_shifts))
        }

    def get_all_employees(self, schedule_data):
        """
        获取所有员工列表

        参数:
            schedule_data (dict): 班表数据

        返回:
            list: 员工姓名列表
        """
        return list(schedule_data.keys())

    def print_employee_summary(self, employee_name, schedule_data):
        """
        打印员工班表摘要

        参数:
            employee_name (str): 员工姓名
            schedule_data (dict): 班表数据
        """
        analysis = self.analyze_employee_schedule(employee_name, schedule_data)

        if not analysis:
            print(f"找不到员工: {employee_name}")
            return

        print("=" * 60)
        print(f"员工: {analysis['name']}")
        print("=" * 60)
        print(f"总天数: {analysis['total_days']} 天")
        print()
        print("班别统计:")
        print("-" * 60)
        for shift_type, count in analysis['stats'].items():
            if count > 0:
                print(f"  {shift_type}: {count} 天")
        print()

    def print_all_summary(self, schedule_data):
        """
        打印所有员工的摘要

        参数:
            schedule_data (dict): 班表数据
        """
        employees = self.get_all_employees(schedule_data)

        print("=" * 80)
        print("班表总览")
        print("=" * 80)
        print(f"总员工数: {len(employees)}")
        print()

        for employee in employees:
            self.print_employee_summary(employee, schedule_data)


def main():
    """主程序示例"""
    # 这是一个使用示例
    # 实际使用时需要提供真实的 Excel 文件路径

    print("班表解析程序")
    print()
    print("使用示例:")
    print("  parser = ScheduleParser('班表.xlsx')")
    print("  schedule_data = parser.parse_schedule(start_row=3, name_col=2)")
    print("  parser.print_employee_summary('员工姓名', schedule_data)")
    print()
    print("说明:")
    print("  - 所有 N 开头的班别 (N, N1, N2, N3) 归类为「夜班」")
    print("  - 所有 M 开头的班别 (M, M1, M2, M3) 归类为「早班」")
    print("  - 所有 A 开头的班别 (A, A1, A2) 归类为「中班」")
    print("  - O: 休息日")
    print("  - P: 休假")
    print("  - BTD: 出差")
    print("  - SL/ML/AL/PL: 病假/休假")


if __name__ == "__main__":
    main()
